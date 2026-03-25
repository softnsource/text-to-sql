# """
# QA Test Runner — hits your local API and appends results back into the xlsx.

# Usage:
#     python qa_runner.py [options]

# Options:
#     --url       Base URL of your API          (default: http://localhost:8000)
#     --session   Session ID to use             (required)
#     --sheet     Which sheet to run            (all | qa | memory, default: all)
#     --input     Path to input xlsx            (default: QA_Results_Full_2026-03-23.xlsx)
#     --output    Path to write results xlsx    (default: QA_Results_<date>.xlsx)
#     --delay     Seconds between requests      (default: 0.5)
#     --timeout   Request timeout in seconds    (default: 60)
#     --only-fail Re-run only previously failed rows
#     --dry-run   Print what would run, no API calls

# Example:
#     python qa_runner.py --url http://localhost:8000 --session YOUR_SESSION_ID
#     python qa_runner.py --session abc123 --sheet memory --delay 1
#     python qa_runner.py --session abc123 --only-fail
# """

# import argparse
# import asyncio
# import copy
# import json
# import sys
# import time
# from datetime import datetime
# from pathlib import Path

# import aiohttp
# import pandas as pd
# from openpyxl import load_workbook
# from openpyxl.styles import Font, PatternFill, Alignment
# from openpyxl.utils import get_column_letter

# # ── Colour palette ────────────────────────────────────────────────────────────
# COLOR = {
#     "pass":    "C6EFCE",   # green
#     "fail":    "FFC7CE",   # red
#     "warn":    "FFEB9C",   # yellow
#     "skip":    "D9D9D9",   # grey
#     "header":  "1F4E79",   # dark blue
#     "header_mem": "375623", # dark green
# }

# FONT_WHITE_BOLD = Font(bold=True, color="FFFFFF", name="Arial", size=10)
# FONT_NORMAL     = Font(name="Arial", size=10)


# # ── Helpers ───────────────────────────────────────────────────────────────────

# def classify_status(status: str) -> str:
#     """Map raw status string → pass / fail / warn / skip."""
#     if not isinstance(status, str):
#         return "skip"
#     s = status.upper()
#     if s.startswith("SQL_OK"):
#         return "pass"
#     if s in ("CLARIFY", "MAYBE_CLARIFY"):
#         return "warn"
#     if s == "TIMEOUT":
#         return "fail"
#     if s.startswith("ERROR") or "FAILED" in s:
#         return "fail"
#     return "warn"


# def cell_color(verdict: str) -> PatternFill:
#     hex_ = COLOR.get(verdict, "FFFFFF")
#     return PatternFill("solid", start_color=hex_, fgColor=hex_)


# def truncate(text, max_len=300):
#     if not isinstance(text, str):
#         return text
#     return text if len(text) <= max_len else text[:max_len] + "…"


# # ── API client ────────────────────────────────────────────────────────────────

# class APIClient:
#     def __init__(self, base_url: str, session_id: str, timeout: int):
#         self.base_url   = base_url.rstrip("/")
#         self.session_id = session_id
#         self.timeout    = aiohttp.ClientTimeout(total=timeout)

#     async def ask(self, question: str) -> dict:
#         payload = {"session_id": self.session_id, "question": question, "page": 1}
#         t0 = time.monotonic()
#         try:
#             async with aiohttp.ClientSession(timeout=self.timeout) as http:
#                 async with http.post(
#                     f"{self.base_url}/api/chat/query",
#                     json=payload,
#                     headers={"Content-Type": "application/json"},
#                 ) as resp:
#                     elapsed = round(time.monotonic() - t0, 2)
#                     body = await resp.json()
#                     return {"ok": True, "elapsed": elapsed, "data": body, "status_code": resp.status}
#         except asyncio.TimeoutError:
#             return {"ok": False, "elapsed": self.timeout.total, "error": "TIMEOUT"}
#         except Exception as e:
#             return {"ok": False, "elapsed": round(time.monotonic() - t0, 2), "error": str(e)}


# def parse_api_result(result: dict) -> tuple[str, str, str]:
#     """
#     Returns (status_str, generated_sql, notes)
#     """
#     if not result["ok"]:
#         err = result.get("error", "unknown")
#         if err == "TIMEOUT":
#             return "TIMEOUT", "", f"Request timed out after {result['elapsed']}s"
#         return f"ERROR: {err}", "", err

#     data = result.get("data", {})
#     mode = data.get("mode", "")
#     sql  = data.get("sql_used", "") or ""
#     text = data.get("text_summary", "") or ""
#     rows = data.get("total_rows", 0) or 0
#     err  = data.get("error", "")

#     if mode == "chat":
#         return "CHAT_RESPONSE", sql, truncate(text)
#     if mode in ("empty", "no_data"):
#         return "NO_DATA", sql, truncate(text)
#     if err:
#         return f"ERROR: {truncate(err, 120)}", sql, truncate(text)
#     if sql:
#         suffix = f" ({rows} rows)" if rows else ""
#         return f"SQL_OK{suffix}", sql, truncate(text, 150)
#     if text:
#         return "CLARIFY", sql, truncate(text)
#     return "UNKNOWN", sql, truncate(str(data))


# # ── Sheet runners ─────────────────────────────────────────────────────────────

# async def run_qa_sheet(df: pd.DataFrame, client: APIClient, delay: float,
#                        only_fail: bool,only_timeout:bool, dry_run: bool) -> pd.DataFrame:
#     results = df.copy()

#     # Add result columns if missing
#     for col in ["Run Status", "Run SQL", "Run Time (s)", "Run Notes", "Run Timestamp"]:
#         if col not in results.columns:
#             results[col] = None

#     total = len(results)
#     for i, row in results.iterrows():
#         q = str(row["Question"]).strip()
#         old_status = str(row.get("Status", "")).strip()

#         # --only-fail filter
#         if only_fail and classify_status(old_status) == "pass":
#             print(f"  [{i+1}/{total}] SKIP (already pass): {q[:60]}")
#             results.at[i, "Run Status"] = "SKIPPED"
#             continue
        
#         if only_timeout and str(row.get("Run Status", "")).strip() != "TIMEOUT":
#             results.at[i, "Run Status"] = row.get("Run Status")  # keep existing
#             continue
#         if dry_run:
#             print(f"  [{i+1}/{total}] DRY-RUN: {q[:80]}")
#             continue

#         print(f"  [{i+1}/{total}] Testing: {q[:80]}", end="", flush=True)
#         result = await client.ask(q)
#         status, sql, notes = parse_api_result(result)
#         elapsed = result.get("elapsed", 0)

#         results.at[i, "Run Status"]    = status
#         results.at[i, "Run SQL"]       = sql
#         results.at[i, "Run Time (s)"]  = elapsed
#         results.at[i, "Run Notes"]     = notes
#         results.at[i, "Run Timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

#         verdict = classify_status(status)
#         icon = "✅" if verdict == "pass" else "⚠️" if verdict == "warn" else "❌"
#         print(f" {icon} {status} ({elapsed}s)")

#         await asyncio.sleep(delay)

#     return results


# async def run_memory_sheet(df: pd.DataFrame, client: APIClient, delay: float,
#                            dry_run: bool , only_timeout=False) -> pd.DataFrame:
#     """
#     Memory tests are grouped conversations — each group resets conversation context.
#     We clear memory between groups by creating a fresh session per group (not possible
#     via API here, so we just log and the user manages sessions manually).
#     Groups are identified by the 'Group' column.
#     """
#     results = df.copy()

#     for col in ["Run Status", "Run SQL", "Run Verdict", "Run Notes", "Run Timestamp"]:
#         if col not in results.columns:
#             results[col] = None

#     # Find groups that have ANY timeout — entire group must re-run
#     if only_timeout:
#         timeout_groups = set(
#             results.loc[results["Run Status"] == "TIMEOUT", "Group"].unique()
#         )
#         if not timeout_groups:
#             print("  No timeout groups found in Memory Tests, skipping.")
#             return results
#         print(f"  Groups with timeouts: {sorted(timeout_groups)} — re-running full groups")

#     total = len(results)
#     current_group = None

#     for i, row in results.iterrows():
#         group = row.get("Group")
#         q     = str(row["Question"]).strip()
#         label = str(row.get("Label", "")).strip()
#         step  = row.get("Step", "?")

#         if group != current_group:
#             current_group = group
#             print(f"\n  ── Group {group}: {label} ──")

#         # Skip groups that don't have timeout issues
#         if only_timeout and group not in timeout_groups:
#             print(f"    Step {step}: SKIP (no timeout in this group)")
#             continue

#         if dry_run:
#             print(f"    Step {step}: DRY-RUN: {q[:80]}")
#             continue

#         print(f"    Step {step}: {q[:80]}", end="", flush=True)
#         result = await client.ask(q)
#         status, sql, notes = parse_api_result(result)
#         elapsed = result.get("elapsed", 0)

#         expected = str(row.get("Expected", "")).strip().lower()
#         memory_verdict = "N/A"
#         if expected:
#             memory_verdict = "PASS" if classify_status(status) == "pass" else "FAIL"

#         results.at[i, "Run Status"]    = status
#         results.at[i, "Run SQL"]       = sql
#         results.at[i, "Run Verdict"]   = memory_verdict
#         results.at[i, "Run Notes"]     = notes
#         results.at[i, "Run Timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

#         icon = "✅" if memory_verdict == "PASS" else "⚠️" if memory_verdict == "N/A" else "❌"
#         print(f" {icon} {status} | verdict={memory_verdict} ({elapsed}s)")

#         await asyncio.sleep(delay)

#     return results


# # ── Summary builder ───────────────────────────────────────────────────────────

# def build_run_summary(qa_df: pd.DataFrame, mem_df: pd.DataFrame) -> dict:
#     def stats(df, status_col):
#         if status_col not in df.columns:
#             return {}
#         statuses = df[status_col].fillna("SKIPPED")
#         total    = len(statuses)
#         passed   = (statuses.str.startswith("SQL_OK") | (statuses == "PASS")).sum()
#         failed   = statuses.str.startswith("ERROR").sum() + (statuses == "TIMEOUT").sum()
#         warn     = statuses.isin(["CLARIFY", "MAYBE_CLARIFY", "NO_DATA", "CHAT_RESPONSE"]).sum()
#         skipped  = (statuses == "SKIPPED").sum()
#         pass_pct = round(passed / max(total - skipped, 1) * 100, 1)
#         return {
#             "total": total, "passed": passed, "failed": failed,
#             "warn": warn, "skipped": skipped, "pass_pct": pass_pct
#         }

#     return {
#         "qa":     stats(qa_df,  "Run Status"),
#         "memory": stats(mem_df, "Run Status"),
#         "run_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
#     }


# # ── Excel writer ──────────────────────────────────────────────────────────────

# def write_results(input_path: str, output_path: str,
#                   qa_df: pd.DataFrame, mem_df: pd.DataFrame, summary: dict):

#     wb = load_workbook(input_path)

#     def style_sheet_header(ws, header_color):
#         for cell in ws[1]:
#             cell.font = FONT_WHITE_BOLD
#             cell.fill = PatternFill("solid", start_color=header_color, fgColor=header_color)
#             cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

#     def auto_width(ws, max_w=60):
#         for col_cells in ws.columns:
#             length = max((len(str(c.value or "")) for c in col_cells), default=10)
#             ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(length + 2, max_w)

#     def write_df_to_sheet(ws, df, status_col, header_color):
#         # Clear existing content
#         for row in ws.iter_rows():
#             for cell in row:
#                 cell.value = None

#         headers = df.columns.tolist()
#         ws.append(headers)
#         style_sheet_header(ws, header_color)

#         for _, row in df.iterrows():
#             ws.append([row.get(h) for h in headers])

#         # Colour rows by status
#         status_idx = headers.index(status_col) + 1 if status_col in headers else None
#         for ws_row in ws.iter_rows(min_row=2):
#             status_val = ws_row[status_idx - 1].value if status_idx else ""
#             verdict    = classify_status(str(status_val or ""))
#             fill       = cell_color(verdict)
#             for cell in ws_row:
#                 cell.font = FONT_NORMAL
#                 if cell.column == status_idx:
#                     cell.fill = fill

#         auto_width(ws)
#         ws.freeze_panes = "A2"

#     # ── All QA Results ──
#     ws_qa = wb["All QA Results"]
#     write_df_to_sheet(ws_qa, qa_df, "Run Status", COLOR["header"])

#     # ── Memory Tests ──
#     ws_mem = wb["Memory Tests"]
#     write_df_to_sheet(ws_mem, mem_df, "Run Status", COLOR["header_mem"])

#     # ── Summary sheet — append a new run block ──
#     ws_sum = wb["Summary"]
#     # Find first empty row
#     last_row = ws_sum.max_row + 2
#     qa_s  = summary["qa"]
#     mem_s = summary["memory"]

#     run_block = [
#         ["── NEW RUN ──", summary["run_at"]],
#         ["QA Tests",       ""],
#         ["  Total",        qa_s.get("total", 0)],
#         ["  Passed",       qa_s.get("passed", 0)],
#         ["  Failed",       qa_s.get("failed", 0)],
#         ["  Warnings",     qa_s.get("warn", 0)],
#         ["  Skipped",      qa_s.get("skipped", 0)],
#         ["  Pass Rate",    f"{qa_s.get('pass_pct', 0)}%"],
#         ["Memory Tests",   ""],
#         ["  Total",        mem_s.get("total", 0)],
#         ["  Passed",       mem_s.get("passed", 0)],
#         ["  Failed",       mem_s.get("failed", 0)],
#         ["  Pass Rate",    f"{mem_s.get('pass_pct', 0)}%"],
#     ]
#     for offset, block_row in enumerate(run_block):
#         r = last_row + offset
#         ws_sum.cell(r, 1, block_row[0])
#         ws_sum.cell(r, 2, block_row[1])
#         if offset == 0:
#             ws_sum.cell(r, 1).font = Font(bold=True, name="Arial", size=11)

#     wb.save(output_path)
#     print(f"\n✅ Results saved → {output_path}")


# # ── CLI entry point ───────────────────────────────────────────────────────────

# async def main():
#     parser = argparse.ArgumentParser(
#         description="QA Test Runner — hit local API and write results to xlsx",
#         formatter_class=argparse.RawDescriptionHelpFormatter,
#         epilog=__doc__,
#     )
#     parser.add_argument("--url",       default="http://localhost:8000", help="Base API URL")
#     parser.add_argument("--session",   required=True,                   help="Session ID")
#     parser.add_argument("--sheet",     default="all",                   choices=["all", "qa", "memory"])
#     parser.add_argument("--input",     default="QA_Results_Full_2026-03-23.xlsx")
#     parser.add_argument("--output",    default=None,                    help="Output xlsx path")
#     parser.add_argument("--delay",     type=float, default=0.5,         help="Delay between requests (s)")
#     parser.add_argument("--timeout",   type=int,   default=60,          help="Request timeout (s)")
#     parser.add_argument("--only-fail", action="store_true",             help="Re-run only failed rows")
#     parser.add_argument("--dry-run",   action="store_true",             help="No API calls, just print")
#     parser.add_argument("--only-timeout", action="store_true", help="Re-run only TIMEOUT rows")
#     args = parser.parse_args()

#     input_path  = Path(args.input)
#     output_path = Path(args.output) if args.output else Path(
#         f"QA_Results_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"
#     )

#     if not input_path.exists():
#         print(f"❌ Input file not found: {input_path}")
#         sys.exit(1)

#     print(f"📂 Input:   {input_path}")
#     print(f"📝 Output:  {output_path}")
#     print(f"🌐 API:     {args.url}")
#     print(f"🔑 Session: {args.session}")
#     print(f"📋 Sheet:   {args.sheet}")
#     print(f"⏱  Delay:   {args.delay}s  |  Timeout: {args.timeout}s")
#     if args.dry_run:
#         print("🔍 DRY RUN — no API calls will be made\n")

#     xl = pd.read_excel(input_path, sheet_name=None)
#     qa_df  = xl.get("All QA Results", pd.DataFrame())
#     mem_df = xl.get("Memory Tests",   pd.DataFrame())

#     client = APIClient(args.url, args.session, args.timeout)

#     # ── Run selected sheets ──
#     if args.sheet in ("all", "qa") and not qa_df.empty:
#         print(f"\n{'='*60}")
#         print(f"Running QA Sheet ({len(qa_df)} tests)…")
#         print(f"{'='*60}")
#         qa_df = await run_qa_sheet(qa_df, client, args.delay, args.only_fail, args.only_timeout, args.dry_run)

#     # if args.sheet in ("all", "memory") and not mem_df.empty:
#     #     print(f"\n{'='*60}")
#     #     print(f"Running Memory Sheet ({len(mem_df)} tests)…")
#     #     print(f"{'='*60}")
#     #     mem_df = await run_memory_sheet(mem_df, client, args.delay, args.dry_run, args.only_timeout)

#     if args.dry_run:
#         print("\n🔍 Dry run complete — no results written.")
#         return

#     # ── Build summary and write ──
#     summary = build_run_summary(qa_df, mem_df)
#     s = summary["qa"]
#     print(f"\n{'='*60}")
#     print(f"QA Results:  {s.get('passed',0)}/{s.get('total',0)} passed  ({s.get('pass_pct',0)}%) "
#           f"| {s.get('failed',0)} failed | {s.get('warn',0)} warnings | {s.get('skipped',0)} skipped")
#     m = summary["memory"]
#     print(f"Memory:      {m.get('passed',0)}/{m.get('total',0)} passed  ({m.get('pass_pct',0)}%)")
#     print(f"{'='*60}")

#     write_results(str(input_path), str(output_path), qa_df, mem_df, summary)


# if __name__ == "__main__":
#     asyncio.run(main())

"""
QA Test Runner — hits your local API and saves results after every question.

Usage:
    python qa_runner.py [options]

Options:
    --url           Base URL of your API          (default: http://localhost:8000)
    --session       Session ID to use             (required)
    --sheet         Which sheet to run            (all | qa | memory, default: all)
    --input         Path to input xlsx            (default: QA_Results_Full_2026-03-23.xlsx)
    --output        Path to write results xlsx    (default: QA_Results_<date>.xlsx)
    --delay         Seconds between requests      (default: 0.5)
    --timeout       Request timeout in seconds    (default: 60)
    --only-fail     Re-run only previously failed rows
    --only-timeout  Re-run only TIMEOUT rows (full group for memory tests)
    --dry-run       Print what would run, no API calls

Example:
    python qa_runner.py --session YOUR_SESSION_ID
    python qa_runner.py --session abc123 --url http://localhost:8000 --timeout 180
    python qa_runner.py --session abc123 --only-timeout --input QA_Results_2026-03-23_14-30.xlsx --output QA_Results_2026-03-23_14-30.xlsx
    python qa_runner.py --session abc123 --only-fail
    python qa_runner.py --session abc123 --sheet memory --delay 1
"""

import argparse
import asyncio
import sys
import time
from datetime import datetime
from pathlib import Path

import aiohttp
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Colour palette ─────────────────────────────────────────────────────────────
COLOR = {
    "pass":       "C6EFCE",
    "fail":       "FFC7CE",
    "warn":       "FFEB9C",
    "skip":       "D9D9D9",
    "header":     "1F4E79",
    "header_mem": "375623",
}

FONT_WHITE_BOLD = Font(bold=True, color="FFFFFF", name="Arial", size=10)
FONT_NORMAL     = Font(name="Arial", size=10)


# ── Helpers ────────────────────────────────────────────────────────────────────

def classify_status(status: str) -> str:
    if not isinstance(status, str):
        return "skip"
    s = status.upper()
    if s.startswith("SQL_OK"):
        return "pass"
    if s in ("CLARIFY", "MAYBE_CLARIFY"):
        return "warn"
    if s == "TIMEOUT":
        return "fail"
    if s.startswith("ERROR") or "FAILED" in s:
        return "fail"
    return "warn"


def cell_color(verdict: str) -> PatternFill:
    hex_ = COLOR.get(verdict, "FFFFFF")
    return PatternFill("solid", start_color=hex_, fgColor=hex_)


def truncate(text, max_len=300):
    if not isinstance(text, str):
        return text
    return text if len(text) <= max_len else text[:max_len] + "..."


# ── API client ─────────────────────────────────────────────────────────────────

class APIClient:
    def __init__(self, base_url: str, session_id: str, timeout: int):
        self.base_url   = base_url.rstrip("/")
        self.session_id = session_id
        self.timeout    = aiohttp.ClientTimeout(total=timeout)

    async def ask(self, question: str) -> dict:
        payload = {"session_id": self.session_id, "question": question, "page": 1}
        t0 = time.monotonic()
        try:
            async with aiohttp.ClientSession(timeout=self.timeout) as http:
                async with http.post(
                    f"{self.base_url}/api/chat/query",
                    json=payload,
                    headers={"Content-Type": "application/json"},
                ) as resp:
                    elapsed = round(time.monotonic() - t0, 2)
                    body    = await resp.json()
                    return {"ok": True, "elapsed": elapsed, "data": body, "status_code": resp.status}
        except asyncio.TimeoutError:
            return {"ok": False, "elapsed": self.timeout.total, "error": "TIMEOUT"}
        except Exception as e:
            return {"ok": False, "elapsed": round(time.monotonic() - t0, 2), "error": str(e)}


def parse_api_result(result: dict) -> tuple:
    if not result["ok"]:
        err = result.get("error", "unknown")
        if err == "TIMEOUT":
            return "TIMEOUT", "", f"Request timed out after {result['elapsed']}s"
        return f"ERROR: {err}", "", err

    data = result.get("data", {})
    mode = data.get("mode", "")
    sql  = data.get("sql_used", "") or ""
    text = data.get("text_summary", "") or ""
    rows = data.get("total_rows", 0) or 0
    err  = data.get("error", "")

    if mode == "chat":
        return "CHAT_RESPONSE", sql, truncate(text)
    if mode in ("empty", "no_data"):
        return "NO_DATA", sql, truncate(text)
    if err:
        return f"ERROR: {truncate(err, 120)}", sql, truncate(text)
    if sql:
        suffix = f" ({rows} rows)" if rows else ""
        return f"SQL_OK{suffix}", sql, truncate(text, 150)
    if text:
        return "CLARIFY", sql, truncate(text)
    return "UNKNOWN", sql, truncate(str(data))


# ── Per-row Excel saver ────────────────────────────────────────────────────────

def save_to_excel(output_path: str, input_path: str,
                  qa_df: pd.DataFrame, mem_df: pd.DataFrame,
                  initialized: bool) -> None:
    """
    Save current state to Excel after every row.
    Only writes Run* columns — all original columns stay untouched.
    """
    base_path = output_path if initialized and Path(output_path).exists() else input_path

    try:
        wb = load_workbook(base_path)
    except Exception as e:
        print(f"\n  Warning: Could not load {base_path}: {e} — using input as fallback")
        wb = load_workbook(input_path)

    def update_sheet(ws, df, run_cols, status_col, header_color):
        if df is None or df.empty:
            return

        # Read existing headers from row 1
        existing_headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

        # Add any missing Run* columns to header row
        for col_name in run_cols:
            if col_name not in existing_headers:
                next_col = len(existing_headers) + 1
                cell = ws.cell(1, next_col, col_name)
                cell.font      = FONT_WHITE_BOLD
                cell.fill      = PatternFill("solid", start_color=header_color, fgColor=header_color)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                existing_headers.append(col_name)

        # Build column index map
        col_map = {name: idx + 1 for idx, name in enumerate(existing_headers) if name}

        # Update only Run* columns row by row
        for excel_row_idx, (_, df_row) in enumerate(df.iterrows(), start=2):
            for col_name in run_cols:
                if col_name not in col_map or col_name not in df_row:
                    continue
                val = df_row[col_name]
                if val is None:
                    continue
                try:
                    if pd.isna(val):
                        continue
                except (TypeError, ValueError):
                    pass
                ws.cell(excel_row_idx, col_map[col_name], val)

            # Colour the Run Status cell
            if status_col in col_map:
                status_val = df_row.get(status_col, "")
                if status_val:
                    verdict = classify_status(str(status_val))
                    ws.cell(excel_row_idx, col_map[status_col]).fill = cell_color(verdict)
                    ws.cell(excel_row_idx, col_map[status_col]).font = FONT_NORMAL

        # Auto column width
        for col_cells in ws.columns:
            length = max((len(str(c.value or "")) for c in col_cells), default=10)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(length + 2, 60)

        ws.freeze_panes = "A2"

    if "All QA Results" in wb.sheetnames and qa_df is not None:
        update_sheet(
            wb["All QA Results"], qa_df,
            run_cols=["Run Status", "Run SQL", "Run Time (s)", "Run Notes", "Run Timestamp"],
            status_col="Run Status",
            header_color=COLOR["header"],
        )

    if "Memory Tests" in wb.sheetnames and mem_df is not None:
        update_sheet(
            wb["Memory Tests"], mem_df,
            run_cols=["Run Status", "Run SQL", "Run Verdict", "Run Notes", "Run Timestamp"],
            status_col="Run Status",
            header_color=COLOR["header_mem"],
        )

    wb.save(output_path)


# ── Sheet runners ──────────────────────────────────────────────────────────────

async def run_qa_sheet(
    df: pd.DataFrame,
    client: APIClient,
    delay: float,
    only_fail: bool,
    only_timeout: bool,
    dry_run: bool,
    input_path: str,
    output_path: str,
    mem_df: pd.DataFrame,
) -> pd.DataFrame:

    results = df.copy()

    for col in ["Run Status", "Run SQL", "Run Time (s)", "Run Notes", "Run Timestamp"]:
        if col not in results.columns:
            results[col] = None

    initialized = Path(output_path).exists()
    total = len(results)

    for idx, (i, row) in enumerate(results.iterrows(), start=1):
        q          = str(row["Question"]).strip()
        old_status = str(row.get("Status", "")).strip()
        run_status = str(row.get("Run Status", "")).strip()

        # --only-fail: skip already passed rows
        if only_fail and classify_status(old_status) == "pass":
            print(f"  [{idx}/{total}] SKIP (already pass): {q[:60]}")
            results.at[i, "Run Status"] = "SKIPPED"
            save_to_excel(output_path, input_path, results, mem_df, initialized)
            initialized = True
            continue

        # --only-timeout: skip rows that are not TIMEOUT
        if only_timeout and run_status != "TIMEOUT":
            continue

        if dry_run:
            print(f"  [{idx}/{total}] DRY-RUN: {q[:80]}")
            continue

        print(f"  [{idx}/{total}] Testing: {q[:80]}", end="", flush=True)
        result  = await client.ask(q)
        status, sql, notes = parse_api_result(result)
        elapsed = result.get("elapsed", 0)

        results.at[i, "Run Status"]    = status
        results.at[i, "Run SQL"]       = sql
        results.at[i, "Run Time (s)"]  = elapsed
        results.at[i, "Run Notes"]     = notes
        results.at[i, "Run Timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        verdict = classify_status(status)
        icon = "✅" if verdict == "pass" else "⚠️" if verdict == "warn" else "❌"
        print(f" {icon} {status} ({elapsed}s)")

        # Save immediately after every question
        save_to_excel(output_path, input_path, results, mem_df, initialized)
        initialized = True

        await asyncio.sleep(delay)

    return results


async def run_memory_sheet(
    df: pd.DataFrame,
    client: APIClient,
    delay: float,
    dry_run: bool,
    only_timeout: bool,
    input_path: str,
    output_path: str,
    qa_df: pd.DataFrame,
) -> pd.DataFrame:

    results = df.copy()

    for col in ["Run Status", "Run SQL", "Run Verdict", "Run Notes", "Run Timestamp"]:
        if col not in results.columns:
            results[col] = None

    # Always define timeout_groups to avoid NameError
    timeout_groups: set = set()
    if only_timeout:
        timeout_groups = set(
            results.loc[results["Run Status"] == "TIMEOUT", "Group"].unique()
        )
        if not timeout_groups:
            print("  No timeout groups found in Memory Tests, skipping.")
            return results
        print(f"  Groups with timeouts: {sorted(timeout_groups)} — re-running full groups")

    initialized   = Path(output_path).exists()
    total         = len(results)
    current_group = None

    for idx, (i, row) in enumerate(results.iterrows(), start=1):
        group = row.get("Group")
        q     = str(row["Question"]).strip()
        label = str(row.get("Label", "")).strip()
        step  = row.get("Step", "?")

        if group != current_group:
            current_group = group
            print(f"\n  -- Group {group}: {label} --")

        # Skip groups without timeout issues
        if only_timeout and group not in timeout_groups:
            if dry_run:
                print(f"    Step {step}: SKIP (no timeout in group {group})")
            continue

        if dry_run:
            print(f"    Step {step}: DRY-RUN: {q[:80]}")
            continue

        print(f"    Step {step}: {q[:80]}", end="", flush=True)
        result  = await client.ask(q)
        status, sql, notes = parse_api_result(result)
        elapsed = result.get("elapsed", 0)

        expected = str(row.get("Expected", "")).strip().lower()
        memory_verdict = "N/A"
        if expected:
            memory_verdict = "PASS" if classify_status(status) == "pass" else "FAIL"

        results.at[i, "Run Status"]    = status
        results.at[i, "Run SQL"]       = sql
        results.at[i, "Run Verdict"]   = memory_verdict
        results.at[i, "Run Notes"]     = notes
        results.at[i, "Run Timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        icon = "✅" if memory_verdict == "PASS" else "⚠️" if memory_verdict == "N/A" else "❌"
        print(f" {icon} {status} | verdict={memory_verdict} ({elapsed}s)")

        # Save immediately after every step
        save_to_excel(output_path, input_path, qa_df, results, initialized)
        initialized = True

        await asyncio.sleep(delay)

    return results


# ── Summary builder ────────────────────────────────────────────────────────────

def build_run_summary(qa_df: pd.DataFrame, mem_df: pd.DataFrame) -> dict:
    def stats(df, status_col):
        if df is None or df.empty or status_col not in df.columns:
            return {}
        statuses = df[status_col].fillna("SKIPPED").astype(str)
        total    = len(statuses)
        passed   = (statuses.str.startswith("SQL_OK") | (statuses == "PASS")).sum()
        failed   = statuses.str.startswith("ERROR").sum() + (statuses == "TIMEOUT").sum()
        warn     = statuses.isin(["CLARIFY", "MAYBE_CLARIFY", "NO_DATA", "CHAT_RESPONSE"]).sum()
        skipped  = (statuses == "SKIPPED").sum()
        pass_pct = round(passed / max(total - skipped, 1) * 100, 1)
        return {
            "total":    total,
            "passed":   int(passed),
            "failed":   int(failed),
            "warn":     int(warn),
            "skipped":  int(skipped),
            "pass_pct": pass_pct,
        }

    return {
        "qa":     stats(qa_df,  "Run Status"),
        "memory": stats(mem_df, "Run Status"),
        "run_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


def append_summary_block(output_path: str, summary: dict) -> None:
    wb       = load_workbook(output_path)
    ws_sum   = wb["Summary"]
    last_row = ws_sum.max_row + 2
    qa_s     = summary["qa"]
    mem_s    = summary["memory"]

    run_block = [
        ["-- NEW RUN --",  summary["run_at"]],
        ["QA Tests",       ""],
        ["  Total",        qa_s.get("total",    0)],
        ["  Passed",       qa_s.get("passed",   0)],
        ["  Failed",       qa_s.get("failed",   0)],
        ["  Warnings",     qa_s.get("warn",     0)],
        ["  Skipped",      qa_s.get("skipped",  0)],
        ["  Pass Rate",    f"{qa_s.get('pass_pct', 0)}%"],
        ["Memory Tests",   ""],
        ["  Total",        mem_s.get("total",   0)],
        ["  Passed",       mem_s.get("passed",  0)],
        ["  Failed",       mem_s.get("failed",  0)],
        ["  Pass Rate",    f"{mem_s.get('pass_pct', 0)}%"],
    ]
    for offset, block_row in enumerate(run_block):
        r = last_row + offset
        ws_sum.cell(r, 1, block_row[0])
        ws_sum.cell(r, 2, block_row[1])
        if offset == 0:
            ws_sum.cell(r, 1).font = Font(bold=True, name="Arial", size=11)

    wb.save(output_path)


# ── CLI entry point ────────────────────────────────────────────────────────────

async def main():
    parser = argparse.ArgumentParser(
        description="QA Test Runner — hit local API, save after every question",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--url",          default="http://localhost:8000", help="Base API URL")
    parser.add_argument("--session",      required=True,                   help="Session ID")
    parser.add_argument("--sheet",        default="all", choices=["all", "qa", "memory"])
    parser.add_argument("--input",        default="QA_Results_Full_2026-03-23_14-30.xlsx")
    parser.add_argument("--output",       default=None,                    help="Output xlsx path")
    parser.add_argument("--delay",        type=float, default=0.5,         help="Delay between requests (s)")
    parser.add_argument("--timeout",      type=int,   default=60,          help="Request timeout (s)")
    parser.add_argument("--only-fail",    action="store_true",             help="Re-run only failed rows")
    parser.add_argument("--only-timeout", action="store_true",             help="Re-run only TIMEOUT rows")
    parser.add_argument("--dry-run",      action="store_true",             help="No API calls, just print")
    args = parser.parse_args()

    input_path  = Path(args.input)
    output_path = Path(args.output) if args.output else Path(
        f"QA_Results_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"
    )

    if not input_path.exists():
        print(f"Input file not found: {input_path}")
        sys.exit(1)

    print(f"Input:        {input_path}")
    print(f"Output:       {output_path}")
    print(f"API:          {args.url}")
    print(f"Session:      {args.session}")
    print(f"Sheet:        {args.sheet}")
    print(f"Delay:        {args.delay}s  |  Timeout: {args.timeout}s")
    print(f"Save mode:    after every question")
    if args.only_timeout:
        print(f"Filter:       TIMEOUT rows only")
    if args.only_fail:
        print(f"Filter:       failed rows only")
    if args.dry_run:
        print("DRY RUN - no API calls will be made\n")

    # Read input file
    xl     = pd.read_excel(input_path, sheet_name=None)
    qa_df  = xl.get("All QA Results", pd.DataFrame())
    mem_df = xl.get("Memory Tests",   pd.DataFrame())

    # Safety check
    if args.sheet in ("all", "qa") and not qa_df.empty and "Question" not in qa_df.columns:
        print(f"'Question' column not found in 'All QA Results'.")
        print(f"Columns found: {qa_df.columns.tolist()}")
        print("Make sure you are using the original input file, not a previously saved output.")
        sys.exit(1)

    client = APIClient(args.url, args.session, args.timeout)

    # ── Run QA sheet ──
    if args.sheet in ("all", "qa") and not qa_df.empty:
        print(f"\n{'='*60}")
        print(f"Running QA Sheet ({len(qa_df)} tests)...")
        print(f"{'='*60}")
        qa_df = await run_qa_sheet(
            qa_df, client, args.delay,
            args.only_fail, args.only_timeout, args.dry_run,
            str(input_path), str(output_path), mem_df,
        )

    # ── Run Memory sheet ──
    if args.sheet in ("all", "memory") and not mem_df.empty:
        print(f"\n{'='*60}")
        print(f"Running Memory Sheet ({len(mem_df)} tests)...")
        print(f"{'='*60}")
        mem_df = await run_memory_sheet(
            mem_df, client, args.delay,
            args.dry_run, args.only_timeout,
            str(input_path), str(output_path), qa_df,
        )

    if args.dry_run:
        print("\nDry run complete - no results written.")
        return

    # ── Append summary block ──
    summary = build_run_summary(qa_df, mem_df)
    s = summary["qa"]
    m = summary["memory"]
    print(f"\n{'='*60}")
    print(f"QA Results:  {s.get('passed',0)}/{s.get('total',0)} passed ({s.get('pass_pct',0)}"
          f"%) | {s.get('failed',0)} failed | {s.get('warn',0)} warnings | {s.get('skipped',0)} skipped")
    print(f"Memory:      {m.get('passed',0)}/{m.get('total',0)} passed ({m.get('pass_pct',0)}%)")
    print(f"{'='*60}")

    if Path(output_path).exists():
        append_summary_block(str(output_path), summary)

    print(f"\nAll done -> {output_path}")


if __name__ == "__main__":
    asyncio.run(main())